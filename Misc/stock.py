from openpyxl import load_workbook, Workbook

# wb = load_workbook('stock.xlsx')
wb = Workbook('stock.xlsx')
sheet = wb.active

# a = sheet.max_column

# while True:
#     part_number = input('Part number: ')
#     for i in range(2, a+1):
#         if part_number == sheet.cell(row=i, column=1).value:
#             query_number = input("Query Number: ")
#             sheet['A1'] = ['query_number']
#         else:
#             print('Part number does not exist')
#
#         print(sheet.cell(row=i, column=5).value)

sheet['A5'] = 'query_number'
wb.save('stock.xlsx')
