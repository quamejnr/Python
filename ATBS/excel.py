import openpyxl

wb = openpyxl.load_workbook('example1.xlsx')
sheet1 = wb['Sheet1']
sheet = wb.active

# for cells in sheet.rows:
#     for values in cells:
#         print(values.value)
#     print('---End of Row---\n')
# print("------THE END-------")
#
# cell = sheet['B1']
# print(cell.value)
# for i in range(1, sheet.max_row + 1):
#     print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=2).value)
#
# for obj in sheet['A1': 'A3']:
#     for cell in obj:
#         print(cell.value)
#     print('---End of Row---')
# print("------THE END-------")
# print(cell.row)
# print(cell.column)
#
# cells = sheet.rows
# for cell in cells:
#     for obj in cell:
#         print(obj.value)
#
# a = sheet.iter_cols(1, 3)
# for i in a:
#     print(i)




